import os 
import json
import openpyxl
from chat_with_model import chat_with_model
from utils import count_char,get_all_file_paths
def ex_q(src_dir,tar_dir,config_dir,content_dir,llm_url, llm_token, llm_method,g_cnt):
    if not os.path.exists(src_dir):
        print('存放用于原始问答对的目录不存在')
        exit()
    if not os.path.exists(tar_dir):
        os.mkdir(tar_dir)
    if llm_url is None or llm_token is None or llm_method is None:
        llm_info = {}
        if os.path.exists(os.path.join(config_dir, 'llm_info.json')):
            with open(os.path.join(config_dir, 'llm_info.json'), 'r') as f:
                llm_info = json.load(f)
        llm_url = llm_info.get('llm_url', '')
        llm_token = llm_info.get('llm_token', '')
        llm_method = llm_info.get('llm_method', '')
    try:
        with open(os.path.join(content_dir, 'ex_q_q_content.txt'), 'r') as f:
            ex_q_content = f.read()
    except:
        print("读取答案生成content失败")
        exit()
    file_path_list = get_all_file_paths(src_dir)
    f_cnt = {}
    for file in file_path_list:
        if os.path.splitext(file)[1] != '.xlsx':
            continue
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active
        out_ws.append(['原始问题']+[cell.value for cell in ws[1]])
        questions_list = []
        for row in ws.iter_rows(2, ws.max_row):
            questions_list.append([cell.value for cell in row][0])
        for row in ws.iter_rows(2, ws.max_row):
            row_data = [cell.value for cell in row]
            cnt = 0
            re_cnt = 0
            while cnt < g_cnt:
                re_cnt += 1
                if re_cnt >= g_cnt*2:
                    break
                try:
                    content = ex_q_content.format(question=row_data[0], answer=row_data[1])
                    for i in range(10):
                        new_q = chat_with_model(llm_url, llm_token, llm_method, content, '请输出一个中文问题')
                        if  count_char(new_q, '?') +count_char(new_q, '？')<= 1:
                            break
                    if  count_char(new_q, '?') +count_char(new_q, '？')> 1:
                        continue
                    if new_q not in questions_list:
                        questions_list.append(new_q)
                        out_ws.append([row_data[0]]+[new_q]+row_data[1:])
                        cnt += 1
                    else:
                        continue
                except:
                    continue
        name = os.path.splitext(os.path.basename(file))[0]
        if name not in f_cnt:
            f_cnt[name] = 0
        out_wb.save(os.path.join(tar_dir, name+'_'+str(f_cnt[name])+'.xlsx'))
        f_cnt[name] += 1
